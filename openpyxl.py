import openpyxl as px
import numpy as np

W = px.load_workbook('filename.xlsx', use_iterators = True)
p = W.get_sheet_by_name(name = 'Sheet1')

a=[]

for row in p.iter_rows():
    for k in row:
        a.append(k.internal_value)

# convert list a to matrix (for example 5*6)
aa= np.resize(a, [5, 6])

# save matrix aa as xlsx file
WW=px.Workbook()
pp=WW.get_active_sheet()
pp.title='NEW_DATA'

f={'A':0,'B':1,'C':2,'D':3,'E':4,'F':5}

#insert values in six columns
for (i,j) in f.items():
    for k in np.arange(1,len(aa)+1):
        pp.cell('%s%d'%(i,k)).value=aa[k-1][j]

WW.save('newfilname.xlsx')
