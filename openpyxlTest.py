"""
Uses openpyxl to generate "numbers.xlsx" (a basic spreadsheet)
"""

import openpyxl as px

WW = px.Workbook()
pp = WW.get_active_sheet()
pp.title = 'SheetNameHere'

f = {'A':1,'B':2,'C':3,'D':4,'E':5,'F':6}

for (i,j) in f.items():
	coord = str(i)+str(1)
	val = "col" + str(i)
	print (coord, " = ", val)
	pp.cell(coord).value = val

for (i,j) in f.items():
    for k in (1, 2, 3):
        #pp.cell('%s%d'%(i,k)).value= k
        coord = str(i)+str(k+1)
        print (coord, " = ", k)
        pp.cell(coord).value = k

WW.save('numbers.xlsx')
