# -*- coding: utf-8 -*-
"""
Created on Sun Sep 30 14:04:27 2018

@author: LALIT ARORA
"""

import sqlite3

def create_connection(db_file):
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except:
        print("Database connection Error!")
        return None
    
    
def columns(dbname,table):
    conn=create_connection(dbname)
    c=conn.cursor()
    conn.commit()
    query="PRAGMA table_info('"+str(table)+"')"
    cols=[]
    for val in c.execute(query):
        cols.append(val[1])
    print (cols)
    
def read_data(dbname):
    conn=create_connection(dbname)
    c=conn.cursor()
    conn.commit()
    tables=[]
    query="select name from sqlite_master where type = 'table'"
    for val in c.execute(query):
        tables=list(val)
    conn.commit()
    d=[]
    for i in range(len(tables)):
        table=tables[i]
        data=[]
        cols=[]
        for val in c.execute("PRAGMA table_info('"+str(table)+"')"):
            cols.append(val[1])
        conn.commit()
        data.append("##".join(cols))
        for val in c.execute("SELECT * from "+str(table)):
            data.append("##".join(list(val)))
        conn.commit()
        d.append(data)
    
    
    for i in range(len(d)):
        temp=d[i]
        a=[]
        for j in range(len(temp)):
            b=list(set(temp[j].split('##')))
            if len(b)==1 and b[0]=="None":
                continue
            else:
                a.append(temp[j])
        d[i]=a
    
    return d
        

def write_workbook(wkb,dbname):
    import openpyxl 
    wb=openpyxl.Workbook()
    wb.save(wkb)
    wb=openpyxl.load_workbook(wkb)
    data=read_data(dbname)
    for i in range(len(data)):
        sheetname="Sheet"+str(i+1)
        wb.create_sheet(sheetname)
        d=data[i]
        for j in range(len(d)):
            d[j]=tuple(d[j].split('##'))
        ac_sheet=wb[sheetname]
        for row in d:
            ac_sheet.append(row)
    wb.save(wkb)
        
if __name__=='__main__':
    import sys
    arguments=sys.argv
    print ("Working..")
    write_workbook(arguments[1],arguments[2])
    print ("Complete..")