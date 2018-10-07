# -*- coding: utf-8 -*-
"""
Created on Mon Sep 17 06:28:21 2018

@author: LALIT ARORA
"""

import sqlite3
import openpyxl

def create_connection(db_file):
    try:
        conn = sqlite3.connect(db_file)
        return conn
    except:
        print("Database connection Error!")
        return None

def u_create_table(wbk,db_file):
    wb=openpyxl.load_workbook(wbk)
    sheets=wb.sheetnames
    for i in range(len(sheets)):
        columns=[]
        ac_sheet=wb[sheets[i]]
        cols = ac_sheet.max_column
        for j in range(1,cols+1):
            temp=(ac_sheet.cell(row=1, column=j).value)
            if temp is not None:
                columns.append("%r"%str(temp))
        print(columns)
        create_table(columns,sheets[i],db_file)
        transfer_data(columns,sheets[i],db_file,wbk)
        
def create_table(cols,sheetname,dbname):
    if len(cols)==0:
        return
    data="CREATE TABLE IF NOT EXISTS "+str(sheetname)+" ( "
    for i in range(len(cols)):
        data=data+str(cols[i])+" text, "
    data=data[:len(data)-2]
    data=data+" );"
    conn=create_connection(dbname)
    c = conn.cursor()
    c.execute(data)
    conn.commit()
    conn.close()
    
def transfer_data(cols,sheetname,dbname,wbk):
    if len(cols)==0:
        return
    wb=openpyxl.load_workbook(wbk)
    ac_sheet=wb[sheetname]
    rows=ac_sheet.max_row
    var=",".join(cols)
    var="INSERT INTO "+str(sheetname)+" ("+var+") VALUES ("
    
    conn=create_connection(dbname)
    c=conn.cursor()
    for i in range(2,rows+1):
        vals=[]
        for j in range(1,len(cols)+1):
            vals.append('"'+str(ac_sheet.cell(row=i, column=j).value)+'"')
        e_vals=",".join(vals)
        e_vals=e_vals+")"
        c.execute(var+e_vals)
        conn.commit()
    conn.close()
        

if __name__ == '__main__':
    import sys
    arguments=sys.argv
    wbk=arguments[1]
    db_file = arguments[2]
    print("Working..")
    u_create_table(wbk,db_file)
    print ("Complete..")
